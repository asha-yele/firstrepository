import openpyxl

filename = "c:\\asha\\UserIdPass.xlsx"
# Load/ open excel file by name
workbook = openpyxl.load_workbook(filename)

# Locate /get perticular sheet
sheet = workbook['UserID']

# Read data from sheet

print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=1, column=2).value)

# Assignment : Read all values from excel

sheet.cell(row=9, column=1).value = "Batch 10"
sheet.cell(row=9, column=2).value = "Batch_10_Pass"

print(sheet.cell(row=9, column=1).value + " " + sheet.cell(row=9, column=2).value)

workbook.save(filename)
