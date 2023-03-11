import openpyxl
import logging

# create and configure logger

logging.basicConfig(filename="CT10.log", format='%(asctime)s %(message)s', filemode='w')
logger = logging.getLogger()

logger.setLevel(logging.DEBUG)

filename = "c:\\UserIDPass.xlsx"
# Load/ open excel file by name
logger.info("Opening workbook : " + filename)
workbook = openpyxl.load_workbook(filename)
logger.debug("Opened workbook : " + filename)


# Locate /get perticular sheet
sheetname = 'UserID'
sheet = workbook[sheetname]
logger.info(f"switched to {sheetname}")

# Read data from sheet
try:
    print(sheet.cell(row=1, column=1).value)
    print(sheet.cell(row=1, column=2).value)
    logger.debug(sheet.cell(row=1, column=1).value + " " + sheet.cell(row=1, column=2).value)
    raise Exception

except Exception as e:
    logger.error("Something went wrong : " , exc_info=e)

# Assignment : Read all values from excel

sheet.cell(row=9, column=1).value = "Batch 10"
sheet.cell(row=9, column=2).value = "Batch_10_Pass"

print(sheet.cell(row=9, column=1).value + " " + sheet.cell(row=9, column=2).value)

workbook.save(filename)