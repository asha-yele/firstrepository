import openpyxl
filename = "c://asha//userpass.xlsx"
workbook = openpyxl.load_workbook(filename)
sheet = workbook["asha"]
print(sheet.cell(row=1, column=2).value)

sheet.cell(row=4, column=2).value = "Batch10"
print(sheet.cell(row=1, column=2).value + " " + sheet.cell(row=4, column=2).value)
workbook.save(filename)
