import openpyxl
filename="c:/asha/excelashafile.xlsx"
# Load/ open excel file by name
workbook = openpyxl.load_workbook(filename)

# wbpath=workbook.path
# print(wbpath)



# Locate /get perticular sheet
sheet = workbook['useridpass']

# Read data from sheet
# print(sheet.cell(row=1,  column=1).value)
# print(sheet.cell(row=1,  column=2).value)
row=1
column=1
for row in sheet:
    for column in sheet:
        if (sheet.cell(row[], column[]).value==''):
            break
            print(sheet.cell(row[] ,column[].value
            row+=1
            column+=1
# print(sheet.cell(row=4, column=1).value + " " + sheet.cell(row=2, column=2).value)




